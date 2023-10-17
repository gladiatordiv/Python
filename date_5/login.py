import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
# Retrieving data from Excel using openpyxl
def reading_excel_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook['Sheet1']

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        email = row[0]
        password = row[1]
        data.append((email, password))
    return data

# Login and Logout
def login_and_logout(driver, email, password):
    driver.get("https://demowebshop.tricentis.com/login")
    email_ = driver.find_element(By.XPATH, "//input[@id='Email']").send_keys(email)
    password_ = driver.find_element(By.XPATH, "//input[@id='Password']").send_keys(password)
    login_button = driver.find_element(By.XPATH, "//input[@value='Log in']")
    login_button.click()
    time.sleep(4)

    try:
        if "Welcome to our store" in driver.page_source:
            print(f"Login successful for {email}")
            # You can add your logout code here
            logout_button = driver.find_element(By.XPATH, "//a[@class='ico-logout']")
            logout_button.click()
            print(f"Logged out for {email}")
            return "Success"
        else:
            print(f"Login failed for {email}")
            return "Failed"
    except Exception as e:
        print(f"Error performing login for {email}: {e}")
        return "Failed"

file_path = "date_5/Book1.xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet = workbook['Sheet1']

data = reading_excel_data(file_path)
driver = webdriver.Chrome()
for row_index, (email, password) in enumerate(data, start=2):
    result = login_and_logout(driver, email, password)
    sheet.cell(row=row_index, column=3, value=result)

try:
    # Close the workbook before saving
    workbook.save("C:/Users/ASUS/OneDrive/Pictures/Book1_updated10.xlsx")
    workbook.close()  
    print("Saved")
except Exception as e:
    print(f"Error in saving: {e}")

driver.quit()
